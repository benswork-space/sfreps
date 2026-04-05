"""
Stage 5: Parse real election results from SF Department of Elections.

Downloads district-level Statement of Vote (dpsov.xlsx) files and extracts
YES/NO vote totals for each local ballot measure by supervisor district.

Data source: https://sfelections.org/results/
"""

import json
import os
import re

import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
RESULTS_DIR = os.path.join(DATA_DIR, "election_results")
PUBLIC_DATA = os.path.join(os.path.dirname(__file__), "..", "public", "data")

# Map election files to election IDs and local measure sheet ranges
ELECTIONS = [
    {
        "file": "nov2024_dpsov.xlsx",
        "election_id": "2024-11",
        "election_name": "November 2024",
    },
    {
        "file": "mar2024_dpsov.xlsx",
        "election_id": "2024-03",
        "election_name": "March 2024",
    },
    {
        "file": "nov2022_dpsov.xlsx",
        "election_id": "2022-11",
        "election_name": "November 2022",
    },
]

# Measure metadata: category + description
MEASURE_INFO = {
    # Nov 2024
    "2024-11-A": {"title": "Affordable Housing Bonds", "category": "housing", "description": "A $300 million bond measure to fund affordable housing construction and preservation, including acquisition of at-risk buildings and new housing for low-income residents."},
    "2024-11-B": {"title": "Community Health & Safety Bonds", "category": "healthcare", "description": "A $390 million bond for community health and safety infrastructure, including hospital upgrades, fire stations, emergency shelters, and family housing facilities."},
    "2024-11-C": {"title": "Inspector General", "category": "government_reform", "description": "Creates an independent Inspector General position to investigate fraud, waste, and abuse in city government, with subpoena power and reporting requirements."},
    "2024-11-D": {"title": "Reduce Commissions, Expand Mayoral Power", "category": "government_reform", "description": "Would have reduced the number of city commissions by roughly half and given the Mayor greater power to appoint and remove commissioners."},
    "2024-11-E": {"title": "Commission Reform Task Force", "category": "government_reform", "description": "Establishes a task force to evaluate the city's commission structure and recommend reforms, as an alternative to the more aggressive Prop D."},
    "2024-11-F": {"title": "Police Deferred Retirement", "category": "public_safety", "description": "Allows police officers eligible for retirement to defer it and continue working for up to 5 years while drawing partial pension benefits, aimed at retaining experienced officers."},
    "2024-11-G": {"title": "Affordable Housing Rental Subsidies", "category": "housing", "description": "Creates the Affordable Housing Opportunity Fund to provide rental subsidies for extremely low-income households, funded at $8.25 million annually."},
    "2024-11-H": {"title": "Firefighter Retirement Age", "category": "first_responders", "description": "Lowers the maximum retirement eligibility age for firefighters from 58 to 55, aligning with neighboring fire departments to improve recruitment."},
    "2024-11-I": {"title": "Nurses & 911 Dispatcher Pensions", "category": "first_responders", "description": "Improves pension benefits for 911 dispatchers and nurses employed by the city to improve recruitment and retention for these critical roles."},
    "2024-11-J": {"title": "Children & Youth Fund Oversight", "category": "education", "description": "Strengthens oversight of the Children and Youth Fund by creating accountability mechanisms for how the city spends money on youth programs."},
    "2024-11-K": {"title": "Close Upper Great Highway", "category": "parks", "description": "Permanently closes the Upper Great Highway to private vehicles, converting it into a public oceanfront recreation space and park."},
    "2024-11-L": {"title": "Rideshare & AV Tax for Muni", "category": "transportation", "description": "Imposes a 1% to 4.5% gross receipts tax on rideshare and autonomous vehicle companies operating in SF, generating ~$25M annually for Muni."},
    "2024-11-M": {"title": "Business Tax Reform", "category": "business_taxes", "description": "Major business tax reform restructuring 14 tax categories into 7, shifting from payroll-based to revenue-based taxation and raising the small business exemption to $5M."},
    "2024-11-N": {"title": "First Responder Student Loan Aid", "category": "first_responders", "description": "Provides student loan repayment assistance for first responders (police, firefighters, paramedics) to attract new recruits amid staffing shortages."},
    "2024-11-O": {"title": "Reproductive Healthcare Access", "category": "healthcare", "description": "Strengthens access to reproductive healthcare including abortion, protects patient and provider privacy, and funds reproductive health services."},
    # Mar 2024
    "2024-03-A": {"title": "Infrastructure Bonds", "category": "transportation", "description": "Infrastructure bonds for city facilities including roads, public transit, and emergency response systems."},
    "2024-03-B": {"title": "Police Minimum Staffing", "category": "public_safety", "description": "Would have increased the mandatory minimum police staffing level from 1,700 to 2,074 officers. Defeated with 67% opposition."},
    "2024-03-C": {"title": "Real Estate Transfer Tax Exemption", "category": "housing", "description": "Exempts commercial-to-residential building conversions from the real estate transfer tax to incentivize creating housing from empty office buildings."},
    "2024-03-D": {"title": "Ethics Law Changes", "category": "government_reform", "description": "Expanded ethics law restrictions on gifts to city officials and strengthened conflict-of-interest rules."},
    "2024-03-E": {"title": "Expanded Police Powers", "category": "public_safety", "description": "Expanded police powers including use of drones, surveillance cameras, and reduced administrative task time for officers."},
    "2024-03-F": {"title": "Drug Screening for Welfare", "category": "drug_policy", "description": "Requires drug screening as a condition of receiving county cash welfare assistance, with mandatory treatment referrals for those who test positive."},
    "2024-03-G": {"title": "Algebra by 8th Grade", "category": "education", "description": "Non-binding policy statement urging SFUSD to offer algebra instruction to all students by eighth grade."},
    # Nov 2022
    "2022-11-A": {"title": "Retiree Benefits Funding", "category": "first_responders", "description": "Adjusts the funding formula for city retiree health and pension benefits to reflect updated actuarial projections."},
    "2022-11-D": {"title": "Affordable Housing Streamlining", "category": "housing", "description": "Would have removed Board of Supervisors approval requirements for certain affordable housing projects to speed up construction."},
    "2022-11-E": {"title": "Housing Project Approval", "category": "housing", "description": "Required Board of Supervisors approval for housing projects in certain areas, as a counter-proposal to Prop D's streamlining approach."},
    "2022-11-F": {"title": "Library Preservation Fund", "category": "libraries", "description": "Renews the Library Preservation Fund for 25 years, providing approximately 97% of the public library system's annual operating budget."},
    "2022-11-I": {"title": "Cars on JFK Drive", "category": "parks", "description": "Would have restored private vehicle access to JFK Drive in Golden Gate Park on weekdays, reversing the pandemic-era car-free policy."},
    "2022-11-J": {"title": "JFK Drive Car-Free", "category": "parks", "description": "Maintains JFK Drive in Golden Gate Park as permanently car-free for pedestrians, cyclists, and recreational use."},
    "2022-11-L": {"title": "Sales Tax for Transportation", "category": "transportation", "description": "Extends a 0.5% city sales tax for 30 years with authorization for up to $1.91 billion in bonds dedicated to transportation improvements."},
    "2022-11-M": {"title": "Vacancy Tax", "category": "business_taxes", "description": "Imposes a tax on owners of residential properties that have been vacant for more than 182 days per year, to discourage housing speculation."},
    "2022-11-N": {"title": "Golden Gate Park Parking", "category": "parks", "description": "Authorizes construction of an underground parking garage in Golden Gate Park to replace surface parking displaced by the JFK Drive closure."},
}


def parse_election_file(filepath, election_id):
    """Parse a dpsov.xlsx file and extract district-level results for local measures."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    results = {}  # measure_key -> {district_num -> {yes, no, total, pct}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue

        # Row 1 (index 1) has the contest name like "MEASURE A" or "PROPOSITION 2"
        contest_name = str(rows[1][0] or "").strip() if len(rows) > 1 and rows[1] else ""

        # Check if this is a local measure (MEASURE A-O) not a state prop
        measure_match = re.match(r'MEASURE\s+([A-Z])\s*$', contest_name, re.IGNORECASE)
        if not measure_match:
            continue

        letter = measure_match.group(1).upper()
        measure_key = "{}-{}".format(election_id, letter)

        if measure_key not in MEASURE_INFO:
            # Unknown measure, skip
            continue

        # Find supervisor district rows (format: "SUP DIST N - Total")
        district_results = {}
        for row in rows:
            if not row or not row[0]:
                continue
            label = str(row[0]).strip()

            # Match "SUP DIST 1 - Total", "SUP DIST 10 - Total", etc.
            dist_match = re.match(r'SUP DIST (\d+) - Total', label)
            if not dist_match:
                continue

            dist_num = int(dist_match.group(1))

            # Columns: [District, RegVoters, Undervotes, _, Overvotes, District2, YES, YES%, NO, NO%, TotalVotes, _]
            try:
                yes_votes = int(row[6]) if row[6] else 0
                no_votes = int(row[8]) if row[8] else 0
                total_votes = int(row[10]) if row[10] else (yes_votes + no_votes)
                yes_pct = round(yes_votes / max(total_votes, 1) * 100, 1)
            except (ValueError, TypeError, IndexError):
                continue

            district_results[dist_num] = {
                "yes": yes_votes,
                "no": no_votes,
                "total": total_votes,
                "support_pct": yes_pct,
            }

        if district_results:
            results[measure_key] = district_results
            info = MEASURE_INFO[measure_key]
            print("  {} (Measure {}): {} - {} districts".format(
                election_id, letter, info["title"], len(district_results)))

    wb.close()
    return results


def fetch_election_results():
    """Parse all election result files and write to district JSON files."""
    print("Processing real election results...\n")

    all_results = {}  # measure_key -> district results

    for election in ELECTIONS:
        filepath = os.path.join(RESULTS_DIR, election["file"])
        if not os.path.exists(filepath):
            print("Missing: {} - download from sfelections.org".format(election["file"]))
            continue

        print("Parsing {}...".format(election["election_name"]))
        results = parse_election_file(filepath, election["election_id"])
        all_results.update(results)

    if not all_results:
        print("\nNo results parsed!")
        return

    # Write to district JSON files
    for dist in range(1, 12):
        dist_path = os.path.join(PUBLIC_DATA, "districts", "district-{}.json".format(dist))
        with open(dist_path) as f:
            dist_data = json.load(f)

        ballot_results = []
        for measure_key, district_results in sorted(all_results.items()):
            if dist not in district_results:
                continue

            dr = district_results[dist]
            info = MEASURE_INFO.get(measure_key, {})
            election_id = "-".join(measure_key.split("-")[:2])
            letter = measure_key.split("-")[-1]

            ballot_results.append({
                "measure_id": "{}-prop-{}".format(election_id, letter.lower()),
                "election": election_id,
                "name": "Proposition {}".format(letter),
                "title": info.get("title", ""),
                "description": info.get("description", ""),
                "category": info.get("category", "other"),
                "support_pct": dr["support_pct"],
                "total_votes": dr["total"],
            })

        # Sort by election (newest first), then by letter
        ballot_results.sort(key=lambda x: (-int(x["election"].replace("-", "")), x["name"]))

        dist_data["ballot_results"] = ballot_results
        with open(dist_path, "w") as f:
            json.dump(dist_data, f, indent=2)

        print("\n  District {}: {} ballot measures".format(dist, len(ballot_results)))

    # Also write a ballot measures index
    measures_index = []
    for measure_key, district_results in sorted(all_results.items()):
        info = MEASURE_INFO.get(measure_key, {})
        election_id = "-".join(measure_key.split("-")[:2])
        letter = measure_key.split("-")[-1]

        # Compute citywide average
        total_yes = sum(dr["yes"] for dr in district_results.values())
        total_all = sum(dr["total"] for dr in district_results.values())
        citywide_pct = round(total_yes / max(total_all, 1) * 100, 1)

        measures_index.append({
            "id": "{}-prop-{}".format(election_id, letter.lower()),
            "election": election_id,
            "name": "Proposition {}".format(letter),
            "title": info.get("title", ""),
            "description": info.get("description", ""),
            "category": info.get("category", "other"),
            "passed": citywide_pct >= 50,
            "citywide_support_pct": citywide_pct,
            "district_results": {str(d): dr["support_pct"] for d, dr in district_results.items()},
        })

    with open(os.path.join(PUBLIC_DATA, "ballot_measures.json"), "w") as f:
        json.dump(measures_index, f, indent=2)

    print("\nWrote {} ballot measures to index".format(len(measures_index)))
    print("Done.")


if __name__ == "__main__":
    fetch_election_results()
